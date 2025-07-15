# -*- coding: utf-8 -*-
"""
修复数据导入问题的脚本
"""

import json
from typing import List, Dict, Any


def parse_csv_line(line: str) -> List[str]:
    """
    更安全的CSV解析器
    """
    if not line:
        return []

    result = []
    current = ''
    inside_quotes = False

    i = 0
    while i < len(line):
        char = line[i]
        next_char = line[i + 1] if i < len(line) - 1 else ''

        if char == '"' and not inside_quotes:
            inside_quotes = True
        elif char == '"' and inside_quotes:
            if next_char == '"':
                current += '"'
                i += 1  # 跳过下一个引号
            else:
                inside_quotes = False
        elif char == ',' and not inside_quotes:
            result.append(current)
            current = ''
        else:
            current += char

        i += 1

    result.append(current)
    return result


def parse_csv(csv_text: str) -> List[Dict[str, Any]]:
    """
    安全的CSV解析
    """
    if not csv_text:
        return []

    # 将CSV文本分割为行
    lines = csv_text.strip().split('\n')
    if len(lines) <= 1:
        print('CSV文件格式无效或为空')
        return []

    # 解析头部
    headers = parse_csv_line(lines[0])

    expected_headers = [
        '电池BT码', 'BMS编号', '电池型号', '循环次数', '返厂原因',
        '返厂时间', '客退地区', '维修状态', '维修项目', '维修费用',
        '维修时间', '快递公司', '运费金额', '责任归属', '维修工时费',
        '原因分析', '改善措施', '维修措施'
    ]

    # 构建字段映射
    header_map = {}
    for index, header in enumerate(headers):
        if not header:
            continue

        # 为每个期望的标题找到最匹配的实际标题
        for expected in expected_headers:
            if expected in header or header in expected:
                header_map[expected] = index
                break

    # 解析数据
    records = []
    for i in range(1, len(lines)):
        if not lines[i].strip():  # 跳过空行
            continue

        # 解析CSV行，处理引号等特殊情况
        values = parse_csv_line(lines[i])
        if not values or len(values) == 0:
            continue

        def safe_get_value(header_name: str) -> str:
            """安全地获取值，防止索引错误"""
            if header_name not in header_map:
                return ''
            index = header_map[header_name]
            if index >= len(values) or values[index] is None:
                return ''
            return str(values[index])

        # 创建记录对象
        import time
        import random
        record = {
            'id': str(int(time.time() * 1000) + random.randint(0, 999)),
            'batteryBtCode': safe_get_value('电池BT码'),
            'bmsNumber': safe_get_value('BMS编号'),
            'batteryModel': safe_get_value('电池型号'),
            'cycleCount': safe_get_value('循环次数'),
            'returnReason': safe_get_value('返厂原因'),
            'returnDate': safe_get_value('返厂时间'),
            'returnArea': safe_get_value('客退地区'),
            'repairStatus': safe_get_value('维修状态'),
            'repairItem': safe_get_value('维修项目'),
            'repairCost': safe_get_value('维修费用'),
            'repairDate': safe_get_value('维修时间'),
            'expressCompany': safe_get_value('快递公司'),
            'shippingCost': safe_get_value('运费金额'),
            'responsibility': safe_get_value('责任归属'),
            'laborCost': safe_get_value('维修工时费'),
            'causeAnalysis': safe_get_value('原因分析'),
            'improvements': safe_get_value('改善措施'),
            'repairMeasures': safe_get_value('维修措施')
        }

        records.append(record)

    return records

def parse_excel(data: bytes) -> List[Dict[str, Any]]:
    """
    安全的Excel解析
    注意：这个函数需要openpyxl库来处理Excel文件
    如果没有安装，请运行: pip install openpyxl
    """
    try:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            print('Excel解析错误: 需要安装openpyxl库')
            print('请运行: pip install openpyxl')
            return []

        import io

        # 从字节数据创建工作簿
        workbook = load_workbook(io.BytesIO(data))
        sheet = workbook.active

        # 获取所有行数据
        raw_data = []
        for row in sheet.iter_rows(values_only=True):
            raw_data.append(list(row))

        if len(raw_data) <= 1:
            raise ValueError('Excel文件格式无效或为空')

        # 转换为CSV文本格式
        csv_lines = []
        for row in raw_data:
            csv_row = []
            for cell in row:
                # 处理单元格内容，包装引号并处理特殊字符
                cell_str = '' if cell is None else str(cell)
                if ',' in cell_str or '"' in cell_str or '\n' in cell_str:
                    cell_str = '"' + cell_str.replace('"', '""') + '"'
                csv_row.append(cell_str)
            csv_lines.append(','.join(csv_row))

        csv_text = '\n'.join(csv_lines)

        # 使用CSV解析器处理
        return parse_csv(csv_text)
    except Exception as error:
        print(f'Excel解析错误: {error}')
        return []

def import_data_from_file(file_path: str) -> List[Dict[str, Any]]:
    """
    从文件导入数据
    """
    try:
        if not file_path:
            print('请提供文件路径')
            return []

        data = []

        # 判断文件类型
        if file_path.lower().endswith('.csv'):
            # 解析CSV
            with open(file_path, 'r', encoding='utf-8') as f:
                csv_text = f.read()
                data = parse_csv(csv_text)
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            # 解析Excel
            with open(file_path, 'rb') as f:
                excel_data = f.read()
                data = parse_excel(excel_data)
        else:
            print('不支持的文件格式，请使用CSV或Excel文件')
            return []

        if not data or len(data) == 0:
            print('没有找到可导入的数据')
            return []

        print(f'成功解析了{len(data)}条记录')
        return data

    except Exception as error:
        print(f'导入处理错误: {error}')
        return []


def save_data_to_json(data: List[Dict[str, Any]], output_file: str = 'data/batteries.json') -> bool:
    """
    将数据保存到JSON文件
    """
    try:
        import os

        # 确保目录存在
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # 如果文件已存在，读取现有数据
        existing_data = []
        if os.path.exists(output_file):
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except:
                existing_data = []

        # 合并数据
        combined_data = existing_data + data

        # 保存到文件
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(combined_data, f, ensure_ascii=False, indent=2)

        print(f'成功保存了{len(combined_data)}条记录到 {output_file}')
        return True

    except Exception as error:
        print(f'保存数据时出错: {error}')
        return False

def main():
    """
    主函数 - 示例用法
    """
    import sys

    if len(sys.argv) < 2:
        print("用法: python fix_import.py <文件路径>")
        print("示例: python fix_import.py data/import.csv")
        return

    file_path = sys.argv[1]

    # 导入数据
    data = import_data_from_file(file_path)

    if data:
        # 保存数据
        success = save_data_to_json(data)
        if success:
            print("数据导入完成！")
        else:
            print("数据保存失败！")
    else:
        print("没有数据可导入")


if __name__ == "__main__":
    main()